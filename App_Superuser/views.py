from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from App_Admin.models import CliPr
from django.db.models import Count
# from weasyprint import HTML
# from io import BytesIO

from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing, Rect, String
import pandas as pd
from io import BytesIO
from django.contrib.staticfiles import finders
#from reportlab.platypus import Image

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Superuser dashboard page rander function:
def superuser1_dashboard(request):
    """
    Renders the superuser dashboard page.
    Extracts cp_id from the URL parameters and passes it to the template.
    """
    cp_id = request.GET.get('cp_id')  # Assuming cp_id is passed as a GET parameter
    email = request.GET.get('email')  # Other parameters can be extracted similarly
    # Fetch client and project data based on cp_id if needed for initial rendering
    context = {
        'cp_id': cp_id,
        'email': email,
        # Add other context variables if necessary
    }
    print(f"check cp_id and email: ::: ::: :{context}")
    return render(request, 'superuser1_dashboard.html', context)

# Table Filtered Data based on cp_id:
def get_filtered_data(request):
    """
    API endpoint to fetch filtered data based on cp_id.
    Expects cp_id as a GET parameter.
    Returns JSON response with the relevant data.
    """
    cp_id = request.GET.get('cp_id')
    status_filter = request.GET.get('status')
    
    if not cp_id:
        return JsonResponse({'error': 'cp_id parameter is missing.'}, status=400)
    
    # Fetch data from the database
    try:
        cp_id_int = int(cp_id)
    except ValueError:
        return JsonResponse({'error': 'Invalid cp_id parameter.'}, status=400)
    
    print(f"check table cp_id : ::: ::: :{cp_id_int}")
    print(f"check table status : ::: ::: :{status_filter}")

    # # Query the CliPr model for entries matching the cp_id
    # client_projects = CliPr.objects.filter(cp_id=cp_id_int)

    # # Query the CliPr model for entries matching the cp_id
    # client_projects = CliPr.objects.filter(status=status).values(
    #     'seeker_name',
    #     'seeker_email',
    #     'provider_name',
    #     'provider_email',
    #     'relationship',
    #     'status'
    # )
    query = CliPr.objects.filter(cp_id=cp_id_int)
   
    if status_filter and status_filter.lower() != "all":  # Apply filter only if it's not "All"
        query = query.filter(status=status_filter)
 
    client_projects = query.values(
        'seeker_name',
        'seeker_email',
        'provider_name',
        'provider_email',
        'relationship',
        'status'
    )
    data = {
        'client_projects': list(client_projects)
    }
    print(data)
    return JsonResponse(data)

# Fetching Overall Dashboard header data: 
def overall_dashboard_header(request):
    """
    API endpoint to fetch overall dashboard header data.
    Returns JSON response with counts of participants, open, in_progress, completed.
    """
    cp_id = request.GET.get('cp_id')
    if not cp_id:
        return JsonResponse({'error': 'cp_id parameter is missing.'}, status=400)
    
    # Fetch data from the database
    try:
        cp_id_int = int(cp_id)
    except ValueError:
        return JsonResponse({'error': 'Invalid cp_id parameter.'}, status=400)
    
    print(f"check dashboard cp_id : ::: ::: :{cp_id_int}")

    queryset = CliPr.objects.filter(cp_id=cp_id_int)

    # Total participants after filtering
    total_participants = queryset.count()

    # Status-wise count
    status_counts = queryset.values('status').annotate(count=Count('status'))

    # Create a dictionary for easy lookup
    status_dict = {item['status']: item['count'] for item in status_counts}
    print(f'Dashboard_status >>>>>>::::::: {status_dict}')
    data = {
        'participants': total_participants,
        'open': status_dict.get('Open', 0),
        'in_progress': status_dict.get('In-Progress', 0),
        'completed': status_dict.get('Completed', 0),
    }
    print(f'Dashboard_Data::::::: {data}')
    return JsonResponse(data)

# Fetch filtered data for PDF generation:
def fetch_filtered_data(cp_id):
    """
    Helper function to fetch filtered data from the database based on cp_id.
    """
    try:
        cp_id_int = int(cp_id)
    except ValueError:
        return {'error': 'Invalid cp_id parameter.'}, 400

    client_projects = CliPr.objects.filter(cp_id=cp_id_int).values(
    
        'project_name',
        'seeker_name',
        'seeker_email',
        'provider_name',
        'provider_email',
        'relationship',
        'status'
    )
    print(f"client_projects : ::: ::: :{client_projects}")
    data = {
        'client_projects': list(client_projects)
    }
    return data, 200    
    
# # Dashboard PDF report generation:

# def superuser_pdf(request):
#     cp_id = request.GET.get('cp_id')
#     if not cp_id:
#         return HttpResponse('cp_id parameter is missing.', status=400)

#     # Fetch data
#     data, status = fetch_filtered_data(cp_id)
#     if status != 200:
#         return JsonResponse(data, status=status)

#     client_projects = data.get('client_projects', [])
#     if not client_projects:
#         return HttpResponse('No data found.', status=404)

#     # Transform data into a pandas DataFrame
#     df = pd.DataFrame(client_projects)
#     print(f" df>>>>>::::::: {df}")
    
#     # Fetch the CliPr object for the given cp_id
#     client_project = CliPr.objects.filter(cp_id=cp_id).first()

#     # Extract client_name and project_name
#     client_name = client_project.client_name
#     project_name = client_project.project_name

#     # Debugging: Print the values
#     print(f"Client Name: {client_name}")
#     print(f"Project Name: {project_name}")

#     # KPIs Calculation
#     participants = len(df)
#     open_count = len(df[df['status'] == 'Open'])
#     in_progress_count = len(df[df['status'] == 'In-Progress'])
#     complete_count = len(df[df['status'] == 'Completed'])

#     # Prepare PDF
#     buffer = BytesIO()
#     pdf_file = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30)
#     story = []

#     # Styles
#     styles = getSampleStyleSheet()
#     header_style = styles['Heading1']
#     header_style.alignment = 1  # Center alignment
#     header_style.fontSize = 14
#     header_style.fontName = 'Helvetica'
#     subheader_style = styles['Heading2']
#     subheader_style.alignment = 1  # Center alignment
#     subheader_style.fontSize = 14
#     subheader_style.fontName = 'Helvetica'
#     normal_style = styles['BodyText']
#     cell_style = styles["BodyText"]
#     cell_style.wordWrap = 'CJK'

#     # Add logo
#     logo_path = finders.find('App_Superuser/image/super_removebg_preview.png')
#     if logo_path:
#         logo = Image(logo_path, width=200, height=50)
#         logo.hAlign = 'CENTER'
#         story.append(logo)
#         story.append(Spacer(1, 16))

#     # Title and KPIs
#     story.append(Spacer(1, 30))  # Adjusted spacing for heading
#     heading_text = f"Overall Dashboard for <b>{client_name }</b> for the Project: <b>{project_name}</b>"
#     story.append(Paragraph(heading_text, header_style))
#     story.append(Spacer(1, -25))
    
#     # Custom KPI display
#     def create_kpi_block(label, value, width=120, height=40, label_color=colors.HexColor("#8d34db"), value_color=colors.HexColor("#f6e7f1")):
#         drawing = Drawing(width, height)
#         # Top section for label
#         drawing.add(Rect(0, height / 2, width, height / 2, fillColor=label_color, strokeColor=label_color))
#         drawing.add(String(width / 2, height * 3 / 4 - 5, label, textAnchor="middle", fontSize=16, fillColor=colors.white))
#         # Bottom section for value
#         drawing.add(Rect(0, 0, width, height / 2, fillColor=value_color, strokeColor=value_color))
#         drawing.add(String(width / 2, height / 4 - 5, str(value), textAnchor="middle", fontSize=16, fillColor=colors.black))
#         return drawing
    
#     kpi_row = [
#         create_kpi_block("Participants", participants),
#         create_kpi_block("Open", open_count),
#         create_kpi_block("In-Progress", in_progress_count),
#         create_kpi_block("Completed", complete_count)
#     ]
    
#     kpi_table = Table([kpi_row], colWidths=[140, 140, 140, 140], rowHeights=[100])
#     kpi_table.setStyle(TableStyle([
#         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
#     ]))
    
#     story.append(kpi_table)
#     story.append(Spacer(1, 24))  # Adjusted spacing
    
#     # Title for details
#     subheading_text = f"Details of <b>{client_name }</b> for the Project: <b>{project_name}</b>"
#     story.append(Paragraph(subheading_text, subheader_style))
#     story.append(Spacer(1, 12))
    
#     # Custom drawing for rounded highlight
#     def highlight_status(text, highlight_color):
#         drawing = Drawing(80, 20)
#         # Draw a rounded rectangle manually
#         drawing.add(Rect(5, 2, 70, 15, fillColor=highlight_color, strokeColor=None))
#         drawing.add(String(40, 7, text, fontSize=10, fillColor=colors.white, textAnchor="middle"))
#         return drawing

#     # Prepare table data
#     data = [
#         ['Seeker Name', 'Seeker Email', 'Provider Name', 'Provider Email', 'Relationship', 'Status']
#     ]
#     for row in df.values.tolist():
#         wrapped_row = []
#         for i, cell in enumerate(row):
#             if i == 5:  # Status column
#                 if cell == 'Open':
#                     wrapped_row.append(highlight_status(cell, colors.HexColor('#28a745')))
#                 elif cell == 'In-Progress':
#                     wrapped_row.append(highlight_status(cell, colors.HexColor('#ffc107')))
#                 elif cell == 'Completed':
#                     wrapped_row.append(highlight_status(cell, colors.HexColor('#007bff')))
#             else:
#                 wrapped_row.append(Paragraph(str(cell), cell_style))
#         data.append(wrapped_row)

#     # Create and style table
#     table = Table(data, colWidths=[80, 120, 80, 120, 80, 80])
#     table.setStyle(TableStyle([
#         ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#8d34db")),
#         ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
#         ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#         ('FONTSIZE', (0, 0), (-1, 0), 10),
#         ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
#         ('GRID', (0, 0), (-1, -1), 1, colors.black),
#     ]))
#     story.append(table)

#     # Build PDF
#     pdf_file.build(story)

#     # Return PDF response
#     buffer.seek(0)
#     response = HttpResponse(buffer, content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename="dashboard_{client_name}_{project_name}.pdf"'
#     return response






def superuser_pdf(request):
    cp_id = request.GET.get('cp_id')
    if not cp_id:
        return HttpResponse('cp_id parameter is missing.', status=400)

    # Fetch data
    data, status = fetch_filtered_data(cp_id)
    if status != 200:
        return JsonResponse(data, status=status)

    client_projects = data.get('client_projects', [])
    if not client_projects:
        return HttpResponse('No data found.', status=404)

    # Transform data into a pandas DataFrame
    df = pd.DataFrame(client_projects)

    # Fetch client_name and project_name                       
    client_project = CliPr.objects.filter(cp_id=cp_id).first()
    if not client_project:
        return HttpResponse('Client or Project not found for the given cp_id.', status=404)

    client_name = client_project.client_name
    project_name = client_project.project_name

    # KPIs Calculation
    participants = len(df)
    open_count = len(df[df['status'] == 'Open'])
    in_progress_count = len(df[df['status'] == 'In-Progress'])
    complete_count = len(df[df['status'] == 'Completed'])

    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    # Add header information
    ws.merge_cells('A1:G1')
    header_cell = ws['A1']
    header_cell.value = f"Dashboard for {client_name} | Project: {project_name}"
    header_cell.font = Font(bold=True, size=14, color='FFFFFF')
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_cell.fill = PatternFill(start_color='8d34db', end_color='8d34db', fill_type='solid')

    # KPIs Section
    kpi_headers = ['Categories', 'Count']
    
    kpi_data = [
        ('Feedback Requested', participants),
        ('Open', open_count),
        ('In-Progress', in_progress_count),
        ('Completed', complete_count),
    ]
    ws.append([])
    ws.append(kpi_headers)
    for row in kpi_data:
        ws.append(row)

    # Apply styles to KPI headers
    for cell in ws[3]:  # Row 3 contains KPI headers
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='f6e7f1', end_color='f6e7f1', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # Table Data Section
    table_headers = ['Project Name','Seeker Name', 'Seeker Email', 'Provider Name', 'Provider Email', 'Relationship', 'Status']
    
    ws.append([])
    ws.append(table_headers)
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    # Style table headers
    for cell in ws[9]:  # Row 9 contains table headers
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='8d34db', end_color='8d34db', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths
    column_widths = [15, 15, 30, 20, 30, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Serve the Excel file as a downloadable response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="dashboard_{client_name}_{project_name}.xlsx"'

    # Save workbook to response
    wb.save(response)

    return response